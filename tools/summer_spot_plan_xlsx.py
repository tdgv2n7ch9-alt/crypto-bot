"""
Пакет 16 -- запись SUMMER_SPOT_PLAN.xlsx через openpyxl. Формат: A4 landscape,
печать на одну страницу по ширине (fitToWidth=1, fitToHeight=0 -- высота
не ужимается, ширина всегда одна страница). Четыре листа: Титул, Ранжир,
Топ-20 DCA, Rug-исключения.
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=16)
SUBTITLE_FONT = Font(size=11, italic=True, color="555555")
WARN_FONT = Font(bold=True, color="C00000")
SECTION_FONT = Font(bold=True, size=13)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")


def _setup_print(ws):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins = PageMargins(left=0.4, right=0.4, top=0.5, bottom=0.5)


def _header_row(ws, row, headers, widths=None):
    for col, text in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=text)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
    if widths:
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w
    # Строковая ссылка, не ws.cell(...) -- вызов .cell() на пустой строке сдвигает
    # внутренний курсор append() на следующую строку, из-за чего первая строка
    # данных после ws.append() уходила пустой (найдено на дымовом прогоне --top 15).
    ws.freeze_panes = f"A{row + 1}"


def _fmt(v, kind="num"):
    if v is None:
        return "н/д"
    if kind == "pct":
        return round(v, 1)
    if kind == "usd":
        return round(v, 2)
    if kind == "usd0":
        return round(v, 0)
    return v


def _write_title_sheet(wb, data):
    ws = wb.active
    ws.title = "Титул"
    _setup_print(ws)
    ws.column_dimensions["A"].width = 100

    r = 1
    ws.cell(row=r, column=1, value="SUMMER SPOT PLAN -- летний спот-ранжир").font = TITLE_FONT
    r += 1
    ws.cell(row=r, column=1,
            value="Ранжир по данным, не гарантия. Решения — владелец.").font = WARN_FONT
    r += 2
    ws.cell(row=r, column=1, value=f"Дата данных: {data['generated_at']}").font = SUBTITLE_FONT
    r += 1
    ws.cell(row=r, column=1,
            value=f"Вселенная: запрошено топ-{data['requested_top_n']} CoinGecko по mcap, "
                  f"получено {data['universe_count']} (стейблы/wrapped исключены).").font = SUBTITLE_FONT
    r += 1
    ws.cell(row=r, column=1,
            value=f"Rug-исключений (скор >= 40): {len(data['rug_excluded'])} -- см. лист «Rug-исключения».").font = SUBTITLE_FONT
    r += 2

    ws.cell(row=r, column=1, value="Источник методологии").font = SECTION_FONT
    r += 1
    for line in [
        "\"Handoff §15\" как отдельный документ НЕ найден в репозитории (проверено предметно --",
        "нет такого раздела ни в одном .md-файле). Методология построена напрямую по критериям,",
        "перечисленным владельцем в задаче Пакета 16 (2026-07-13) -- владелец подтвердил этот",
        "источник явно (\"работать по критериям из вашего сообщения\").",
    ]:
        ws.cell(row=r, column=1, value=line)
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="Как считается скор (прозрачная аддитивная формула, база 50)").font = SECTION_FONT
    r += 1
    for line in [
        "+12  Просадка от ATH 40-90% (зона интереса, не экстремум)      -6  Просадка <15% (не дно)",
        "+8   7д импульс 0..+12% (ранний отскок)                        -4  7д импульс >+12% (риск догонять)",
        "+6   30д отрицательный (в коррекции, по сценарию)              -3  30д уже растёт",
        "+8   90д < -20% (продолжительная коррекция)                    -4  90д растёт",
        "+10  VRVP: цена внутри зоны макс. объёма 180д (накопление)",
        "+6   MCap/TVL < 3x (капа не оторвана от реального использования)  -4  MCap/TVL > 15x",
        "+5   реальная выручка протокола за 30д (DeFiLlama)",
        "-rug_score*0.3  штраф пропорционально rug-риску (0-39, монеты >=40 исключены полностью)",
        "+5   FDV/MCap <= 1.3x (малый навес эмиссии)                    -8  FDV/MCap >= 3x",
        "+5   Объём/MCap 1-15% (здоровая ликвидность)             -5/-3  слишком низкий/высокий",
        "+6   активная LONG-зона Королева (journal/watch_zones.json)",
    ]:
        ws.cell(row=r, column=1, value=line)
        r += 1
    r += 1

    ws.cell(row=r, column=1, value="Сценарий владельца").font = SECTION_FONT
    r += 1
    ws.cell(row=r, column=1,
            value="Отскок -> плавное снижение -> летнее дно альтов. Формула сознательно contrarian: "
                  "просадка/коррекция оценивается ВЫШЕ (это ранжир для набора позиции у дна, "
                  "не momentum-скринер прорывов).")
    ws.cell(row=r, column=1).alignment = WRAP
    ws.row_dimensions[r].height = 30
    r += 2

    ws.cell(row=r, column=1, value="Ярусы (целевые доли портфеля, назначены владельцем)").font = SECTION_FONT
    r += 1
    ws.cell(row=r, column=1, value="Мейджоры 50%: BTC, ETH, SOL")
    r += 1
    ws.cell(row=r, column=1, value="Качество 35%: AAVE, UNI, LINK, MORPHO, ENA")
    r += 1
    ws.cell(row=r, column=1, value="Бета 15%: SUI, AVAX, WLD, JASMY + прочие по общему скору")
    r += 2

    ws.cell(row=r, column=1, value="Честные ограничения данных").font = SECTION_FONT
    r += 1
    for line in [
        "-- 90д импульс и VRVP-эвристика -- только для монет с парой на Binance (не все топ-150 имеют).",
        "-- TVL/выручка -- только для DeFi-протоколов, найденных в DeFiLlama по совпадению symbol/name;",
        "   для L1-монет (BTC/ETH/SOL/AVAX/SUI/...) и утилити-токенов честно \"н/д (не применимо)\".",
        "-- Rug-скор считается БЕЗ доп. CoinGecko-запросов на монету (см. RUG_WATCHLIST.md -- прошлая",
        "   попытка так и упала в 429 почти на всех монетах) -- часть детекторов честно недоступна,",
        "   не влияет на max_possible_score.",
        "-- \"х100-кандидаты\" в бета-ярусе -- топ-скорящие монеты этого же ранжира вне уже названных",
        "   ярусов, НЕ отдельный запуск живого /x100-сканера бота.",
    ]:
        ws.cell(row=r, column=1, value=line)
        r += 1


def _write_ranking_sheet(wb, data):
    ws = wb.create_sheet("Ранжир")
    _setup_print(ws)
    headers = ["Ранг", "Символ", "Название", "Ярус", "Скор", "Цена $", "MCap $",
               "ATH %", "7д %", "30д %", "90д %", "Rug", "FDV/MCap", "Vol/MCap %",
               "TVL $", "Выручка30д $", "Зона Королева", "Топ-факторы"]
    widths = [6, 9, 16, 9, 7, 12, 15, 8, 7, 7, 7, 6, 9, 10, 14, 13, 8, 60]
    _header_row(ws, 1, headers, widths)

    for r in data["results"]:
        c = r["coin"]
        prof = r["profile"]
        tvl = r["tvl_rev"]
        ch90 = prof.get("ch_90d") if prof.get("ok") else None
        top_factors = sorted(r["factors"], key=lambda f: -abs(f[1]))[:3]
        factors_str = "; ".join(f"{lbl} ({d:+.1f})" for lbl, d in top_factors if d != 0)
        row = [
            r["final_rank"], c["symbol"], c["name"], r["tier"], r["score"],
            _fmt(c["price"], "usd"), _fmt(c["market_cap"], "usd0"),
            _fmt(c["ath_change_pct"], "pct"), _fmt(c["ch_7d"], "pct"),
            _fmt(c["ch_30d"], "pct"), _fmt(ch90, "pct"),
            r["rug"].get("score", 0),
            _fmt(c["fdv"] / c["market_cap"], "usd") if c.get("fdv") and c["market_cap"] else "н/д",
            _fmt(c["volume_24h"] / c["market_cap"] * 100, "pct") if c["market_cap"] else "н/д",
            _fmt(tvl.get("tvl_usd"), "usd0"),
            _fmt(tvl.get("revenue_30d_usd"), "usd0"),
            "да" if r["korolev"] else "",
            factors_str,
        ]
        ws.append(row)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[-1].alignment = WRAP


def _write_top20_sheet(wb, data):
    ws = wb.create_sheet("Топ-20 DCA")
    _setup_print(ws)
    headers = ["Ранг", "Символ", "Ярус", "Скор", "Зона lo", "Зона hi",
               "Вход 1 (50%)", "Вход 2 (30%)", "Вход 3 (20%)", "SL", "TP",
               "Инвалидация", "Источник лестницы", "Заметка"]
    widths = [6, 9, 9, 7, 11, 11, 13, 13, 13, 11, 11, 30, 34, 55]
    _header_row(ws, 1, headers, widths)

    for r in data["top20"]:
        c = r["coin"]
        li = r["ladder_info"]
        ladder = li.get("ladder") or []
        prices = [None, None, None]
        for i, lvl in enumerate(ladder[:3]):
            prices[i] = lvl.get("price")
        zone = li.get("zone") or {}
        row = [
            r["final_rank"], c["symbol"], r["tier"], r["score"],
            _fmt(zone.get("lo"), "usd"), _fmt(zone.get("hi"), "usd"),
            _fmt(prices[0], "usd"), _fmt(prices[1], "usd"), _fmt(prices[2], "usd"),
            _fmt(li.get("sl"), "usd"), _fmt(li.get("tp"), "usd"),
            li.get("invalidation") or "н/д",
            li.get("source") or "н/д",
            li.get("note") or "",
        ]
        ws.append(row)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[11].alignment = WRAP
        row[13].alignment = WRAP


def _write_rug_excluded_sheet(wb, data):
    ws = wb.create_sheet("Rug-исключения")
    _setup_print(ws)
    headers = ["Символ", "Название", "Rug-скор", "Причины"]
    widths = [10, 20, 10, 90]
    _header_row(ws, 1, headers, widths)

    for ex in data["rug_excluded"]:
        ws.append([ex["symbol"], ex["name"], ex["score"], "; ".join(ex["reasons"])])

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[-1].alignment = WRAP


def write_workbook(data: dict, out_path: str):
    wb = Workbook()
    _write_title_sheet(wb, data)
    _write_ranking_sheet(wb, data)
    _write_top20_sheet(wb, data)
    _write_rug_excluded_sheet(wb, data)
    wb.save(out_path)
